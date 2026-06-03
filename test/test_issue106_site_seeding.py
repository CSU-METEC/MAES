"""Tests for GitHub issue CSU-METEC/MAES#106 — identical per-site parquet output.

The bug: ``runSim`` seeds the simulation RNG from the MC run number alone
(``SimRNG.seed(mcRunNum)``, introduced by the #69 RNG consolidation). The seed
contains no site identity, so two sites with identical study definitions replay
bit-for-bit identical random streams and produce numerically identical
SiteSummary output — the reporter's symptom.

The fix: compose the seed from ``[baseSeed?, crc32(siteName), mcRunNum]``
(``SimRNG.composeSeed``). Site identity separates sites within a run; the MC
run number keeps per-run distinctness; the optional ``--randomSeed`` base keeps
whole-simulation reproducibility (#96).

Two tiers:

* Fast unit tests of the seed-composition contract (no engine).
* One slow engine tier (``@pytest.mark.slow``) reproducing the issue exactly:
  two byte-identical study sheets run via ``--directory``, asserting their
  SiteSummary numbers differ, that a default rerun is bit-identical
  (preserving #69's per-run determinism), and that ``--randomSeed`` shifts
  both sites while keeping them distinct.

Run the slow tier in the engine's runtime env (MAES conda env).
"""

import pathlib
import shutil
import subprocess
import sys
import zlib
from dataclasses import dataclass

import pandas as pd
import pytest

sys.path.insert(0, str(pathlib.Path(__file__).parent.parent / "src"))

import SimRNG

MAES_ROOT = pathlib.Path(__file__).parent.parent
SOURCE_STUDY = MAES_ROOT / "input/Studies/C3/C3_Prototypical_Sites/P1_1stage_noflare.xlsx"
FIXTURE_SIM_DAYS = 7
MC_ITERATIONS = 1


# --------------------------------------------------------------------------- fast: seed contract

def test_compose_seed_distinct_sites_distinct_seeds():
    """Two sites at the same MC run must get different seed material (#106)."""
    assert SimRNG.composeSeed(0, siteName="SiteA") != SimRNG.composeSeed(0, siteName="SiteB")


def test_compose_seed_distinct_mc_runs_distinct_seeds():
    """Same site, different MC runs must stay distinct (#69 contract)."""
    assert SimRNG.composeSeed(0, siteName="SiteA") != SimRNG.composeSeed(1, siteName="SiteA")


def test_compose_seed_deterministic():
    """Same (site, mcRun, baseSeed) is reproducible across calls."""
    assert SimRNG.composeSeed(3, siteName="X", baseSeed=42) == SimRNG.composeSeed(3, siteName="X", baseSeed=42)


def test_compose_seed_base_seed_prepended():
    """--randomSeed contributes a distinct leading component (#96), seed 0 included."""
    noBase = SimRNG.composeSeed(1, siteName="X")
    withBase = SimRNG.composeSeed(1, siteName="X", baseSeed=0)
    assert withBase != noBase
    assert withBase[0] == 0


def test_compose_seed_site_key_is_crc32():
    """Site component is crc32 of the site name — stable across platforms/runs."""
    seed = SimRNG.composeSeed(5, siteName="SiteA")
    assert seed == [zlib.crc32(b"SiteA"), 5]


def test_compose_seed_legacy_without_site():
    """No site, no base seed → just the MC run number (legacy #69 behavior)."""
    assert SimRNG.composeSeed(7) == [7]


def test_streams_differ_between_sites():
    """End-to-end through the generator: identical draws were the #106 bug."""
    SimRNG.seed(SimRNG.composeSeed(0, siteName="SiteA"))
    drawsA = list(map(lambda _: SimRNG.random(), range(100)))
    SimRNG.seed(SimRNG.composeSeed(0, siteName="SiteB"))
    drawsB = list(map(lambda _: SimRNG.random(), range(100)))
    assert drawsA != drawsB


# --------------------------------------------------------------------------- slow: engine repro

@dataclass
class EngineRun:
    siteA: pd.DataFrame
    siteB: pd.DataFrame


def _makeFixtureSheets(studiesDir: pathlib.Path) -> None:
    """Write SiteA.xlsx / SiteB.xlsx — byte-identical short-duration copies of P1.

    P1's only formula cells (three ``=76.68…/10`` constants on Cycling Wells) are
    materialized to literals: openpyxl does not carry cached formula results, so
    a resave would otherwise leave NaNs where the engine expects numbers.
    """
    import openpyxl

    wb = openpyxl.load_workbook(SOURCE_STUDY)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    expr = cell.value.lstrip("=")
                    num, denom = expr.split("/")
                    cell.value = float(num) / float(denom)
    params = wb["Global Simulation Parameters"]
    for row in params.iter_rows(min_col=1, max_col=2):
        if row[0].value == "Simulation Duration [Days]":
            row[1].value = FIXTURE_SIM_DAYS
    siteA = studiesDir / "SiteA.xlsx"
    wb.save(siteA)
    shutil.copyfile(siteA, studiesDir / "SiteB.xlsx")


def _makeInputRoot(tmpRoot: pathlib.Path) -> pathlib.Path:
    """Build an inputRoot whose Studies/ holds only the fixture pair.

    CuratedData and ModelFormulation are symlinked from the repo when possible
    (Linux/mac), copied otherwise (Windows without symlink privilege).
    """
    inputRoot = tmpRoot / "input"
    studiesDir = inputRoot / "Studies" / "Issue106"
    studiesDir.mkdir(parents=True)
    for sub in ("CuratedData", "ModelFormulation"):
        src, dst = MAES_ROOT / "input" / sub, inputRoot / sub
        try:
            dst.symlink_to(src, target_is_directory=True)
        except OSError:
            shutil.copytree(src, dst)
    _makeFixtureSheets(studiesDir)
    return inputRoot


def _runEngine(inputRoot: pathlib.Path, outDir: pathlib.Path, seed=None) -> EngineRun:
    """Run the two-site study; return each site's SiteSummary (site column dropped)."""
    outDir.mkdir(parents=True, exist_ok=True)
    cmd = [
        sys.executable, "src/SiteMain2.py",
        "-i", str(inputRoot),
        "-or", str(outDir),
        "-dr", "Issue106",
        # getConfig reads the -s study for global params even in -dr mode, and the
        # default (MEET2/ConstantSeparator.xlsx) does not exist in this inputRoot.
        "-s", "Issue106/SiteA.xlsx",
        "-mc", str(MC_ITERATIONS),
    ]
    if seed is not None:
        cmd += ["-rs", str(seed)]
    r = subprocess.run(cmd, cwd=MAES_ROOT, capture_output=True, text=True, timeout=1800)
    assert r.returncode == 0, f"engine failed (seed={seed}).\nSTDOUT:\n{r.stdout}\nSTDERR:\n{r.stderr}"

    datasets = list(filter(lambda d: d.is_dir(), outDir.rglob("SiteSummary")))
    assert len(datasets) == 1, f"expected one SiteSummary dataset under {outDir}, found {datasets}"
    return EngineRun(
        siteA=_loadSite(datasets[0], "SiteA"),
        siteB=_loadSite(datasets[0], "SiteB"),
    )


def _loadSite(dataset: pathlib.Path, site: str) -> pd.DataFrame:
    df = pd.read_parquet(dataset, filters=[("site", "=", site)])
    assert not df.empty, f"no SiteSummary rows for {site}"
    sortCols = list(filter(lambda c: df[c].dtype == object and c != "readings", df.columns))
    return (df.drop(columns=["site", "readings"], errors="ignore")
              .sort_values(sortCols)
              .reset_index(drop=True))


def _framesEqual(a: pd.DataFrame, b: pd.DataFrame) -> bool:
    return a.shape == b.shape and a.equals(b)


@pytest.fixture(scope="module")
def engineRuns(tmp_path_factory):
    tmpRoot = tmp_path_factory.mktemp("issue106")
    inputRoot = _makeInputRoot(tmpRoot)
    return {
        "default": _runEngine(inputRoot, tmpRoot / "out_default"),
        "default_repeat": _runEngine(inputRoot, tmpRoot / "out_default_repeat"),
        "seed42": _runEngine(inputRoot, tmpRoot / "out_seed42", seed=42),
    }


@pytest.mark.slow
def test_identical_sites_produce_distinct_output(engineRuns):
    """THE #106 regression: byte-identical sheets must not yield identical numbers."""
    run = engineRuns["default"]
    assert not _framesEqual(run.siteA, run.siteB), (
        "SiteA and SiteB SiteSummary outputs are numerically identical — "
        "the RNG seed contains no site identity (CSU-METEC/MAES#106)"
    )


@pytest.mark.slow
def test_default_rerun_is_deterministic(engineRuns):
    """#69's contract survives the fix: same inputs, same seeds → identical output."""
    first, second = engineRuns["default"], engineRuns["default_repeat"]
    assert _framesEqual(first.siteA, second.siteA)
    assert _framesEqual(first.siteB, second.siteB)


@pytest.mark.slow
def test_random_seed_shifts_both_sites_and_keeps_them_distinct(engineRuns):
    """--randomSeed (#96) composes with site identity instead of replacing it."""
    default, seeded = engineRuns["default"], engineRuns["seed42"]
    assert not _framesEqual(seeded.siteA, seeded.siteB)
    assert not _framesEqual(default.siteA, seeded.siteA)
    assert not _framesEqual(default.siteB, seeded.siteB)
