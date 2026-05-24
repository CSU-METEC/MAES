"""Input-root split — fast tests (curatedRoot / studyInputRoot overlay) — CSU-METEC/MAES #93.

No engine: the resolveInputRef resolution matrix and the getConfig study-file overlay + root-alias
precedence. The slow, engine-driven Output-Equivalence cases live in
test_input_root_split_integration.py.
"""

import os
import sys
import shutil
import pathlib

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import AppUtils as au

MAES_ROOT = pathlib.Path(__file__).parent.parent
STUDY = "C3/C3_Prototypical_Sites/P1_1stage_noflare.xlsx"


def _stage_study(root: pathlib.Path) -> pathlib.Path:
    """Copy the curated study into root/Studies/<STUDY> so an overlay can serve it."""
    dst = root / "Studies" / STUDY
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(MAES_ROOT / "input" / "Studies" / STUDY, dst)
    return dst


def _stage_study_with_factor_override(root: pathlib.Path, factor_ref: str) -> pathlib.Path:
    """Stage the study in the overlay with 'Activity / Emission Factor Name' set to factor_ref,
    so getConfig picks up a study-provided factors override (exercises the C1 resolution hook)."""
    import openpyxl
    dst = root / "Studies" / STUDY
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(MAES_ROOT / "input" / "Studies" / STUDY)
    ws = wb["Global Simulation Parameters"]
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Activity / Emission Factor Name":
            ws.cell(row=row, column=2, value=factor_ref)
            break
    wb.save(dst)
    return dst


# --------------------------------------------------------------------------- resolveInputRef unit

def test_resolve_noop_when_roots_equal():
    assert str(au.resolveInputRef("input/Studies/x.xlsx", "input", "input")) == "input/Studies/x.xlsx"


def test_resolve_absolute_path_bypasses():
    assert str(au.resolveInputRef("/etc/hostname", "input", "/overlay")) == "/etc/hostname"


def test_resolve_bare_ref_overlay_present(tmp_path):
    (tmp_path / "gc").mkdir()
    (tmp_path / "gc" / "x.csv").write_text("a")
    assert str(au.resolveInputRef("gc/x.csv", "input", str(tmp_path))) == str(tmp_path / "gc" / "x.csv")


def test_resolve_bare_ref_overlay_absent_falls_through(tmp_path):
    assert str(au.resolveInputRef("gc/missing.csv", "input", str(tmp_path))) == "gc/missing.csv"


def test_resolve_templated_ref_overlay_present(tmp_path):
    (tmp_path / "Studies").mkdir()
    (tmp_path / "Studies" / "s.xlsx").write_text("a")
    got = au.resolveInputRef("input/Studies/s.xlsx", "input", str(tmp_path))
    assert str(got) == str(tmp_path / "Studies" / "s.xlsx")


def test_resolve_templated_ref_overlay_absent_keeps_curated(tmp_path):
    assert str(au.resolveInputRef("input/Studies/none.xlsx", "input", str(tmp_path))) == "input/Studies/none.xlsx"


# --------------------------------------------------------------------------- getConfig overlay (fast)

@pytest.fixture
def at_maes_root(monkeypatch):
    monkeypatch.chdir(MAES_ROOT)


def test_getconfig_overlay_picks_studyinputroot(at_maes_root, tmp_path):
    _stage_study(tmp_path)
    cm, _ = au.getConfig(commandArgs=["-s", STUDY, "-sir", str(tmp_path)])
    assert cm.getConfigVar("studyFilename").startswith(str(tmp_path))


def test_getconfig_overlay_absent_falls_back_to_curated(at_maes_root, tmp_path):
    cm, _ = au.getConfig(commandArgs=["-s", STUDY, "-sir", str(tmp_path)])
    assert cm.getConfigVar("studyFilename") == f"input/Studies/{STUDY}"


def test_getconfig_default_is_single_root(at_maes_root):
    cm, _ = au.getConfig(commandArgs=["-s", STUDY])
    assert cm.getConfigVar("curatedRoot") == "input"
    assert cm.getConfigVar("studyInputRoot") == "input"
    assert cm.getConfigVar("studyFilename") == f"input/Studies/{STUDY}"


def test_getconfig_inputroot_alias_seeds_both(at_maes_root, tmp_path):
    _stage_study(tmp_path)
    cm, _ = au.getConfig(commandArgs=["-s", STUDY, "-i", str(tmp_path)])
    assert cm.getConfigVar("curatedRoot") == str(tmp_path)
    assert cm.getConfigVar("studyInputRoot") == str(tmp_path)


def test_getconfig_explicit_curated_and_studyinput(at_maes_root, tmp_path):
    cm, _ = au.getConfig(commandArgs=["-s", STUDY, "-cr", "input", "-sir", str(tmp_path)])
    assert cm.getConfigVar("curatedRoot") == "input"
    assert cm.getConfigVar("studyInputRoot") == str(tmp_path)


FACTOR_OVERRIDE_REF = "CuratedData/FactorsFileReference/OverrideFactors.csv"


def test_getconfig_factor_override_resolves_to_studyinputroot(at_maes_root, tmp_path):
    """A study-provided factors override is overlay-resolved to studyInputRoot when present (C1)."""
    _stage_study_with_factor_override(tmp_path, FACTOR_OVERRIDE_REF)
    target = tmp_path / FACTOR_OVERRIDE_REF
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("MajorEquipment\n")  # existence is enough for path resolution
    cm, _ = au.getConfig(commandArgs=["-s", STUDY, "-sir", str(tmp_path)])
    assert cm.getConfigVar("factorName") == str(target)


def test_getconfig_factor_override_absent_keeps_ref(at_maes_root, tmp_path):
    """When the overlay lacks the override file, factorName is left unchanged (falls through)."""
    _stage_study_with_factor_override(tmp_path, FACTOR_OVERRIDE_REF)
    cm, _ = au.getConfig(commandArgs=["-s", STUDY, "-sir", str(tmp_path)])
    assert cm.getConfigVar("factorName") == FACTOR_OVERRIDE_REF
