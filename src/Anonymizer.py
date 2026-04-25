import json
import logging
import random
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl

import BundleFormat as bf

logger = logging.getLogger(__name__)

_KEY_VERSION = '1.0'


class _AnonState:
    """Accumulates anonymization mappings and generates consistent anonymized values."""

    def __init__(self):
        self.facilityMap: dict[str, str] = {}
        self.unitMap:     dict[str, str] = {}
        self.studyMap:    dict[str, str] = {}
        self.latOffset = random.uniform(-5.0, 5.0)
        self.lonOffset = random.uniform(-5.0, 5.0)

    def mapFacility(self, val: str) -> str:
        if val not in self.facilityMap:
            self.facilityMap[val] = f"facility_{len(self.facilityMap) + 1:03d}"
        return self.facilityMap[val]

    def mapUnit(self, val: str) -> str:
        if val not in self.unitMap:
            self.unitMap[val] = f"unit_{len(self.unitMap) + 1:03d}"
        return self.unitMap[val]

    def mapStudy(self, name: str) -> str:
        if name not in self.studyMap:
            self.studyMap[name] = f"study_{len(self.studyMap) + 1:03d}"
        return self.studyMap[name]

    def toKeyDict(self) -> dict:
        return {
            'version':          _KEY_VERSION,
            'coordinateOffset': {'lat': self.latOffset, 'lon': self.lonOffset},
            'studies':          dict(self.studyMap),
            'facilityIDs':      dict(self.facilityMap),
            'unitIDs':          dict(self.unitMap),
        }


def _colRole(header: str) -> str | None:
    """Return the anonymization role for a column header, or None if not a target."""
    if header is None:
        return None
    h = str(header).strip()
    if h == 'Facility ID' or h.endswith('FacilityID'):
        return 'facilityID'
    if h == 'Unit ID' or h.endswith('UnitID'):
        return 'unitID'
    if h == 'Latitude':
        return 'lat'
    if h == 'Longitude':
        return 'lon'
    return None


def _processSheet(ws, *, facilityFn, unitFn, latFn, lonFn) -> None:
    """Apply per-role transform functions to all data cells in a sheet."""
    headerRow = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if headerRow is None:
        return

    colRoles = {
        colIdx: role
        for colIdx, header in enumerate(headerRow, start=1)
        if (role := _colRole(header)) is not None
    }
    if not colRoles:
        return

    fns = {'facilityID': facilityFn, 'unitID': unitFn, 'lat': latFn, 'lon': lonFn}

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            role = colRoles.get(cell.column)
            if role is None or cell.value is None:
                continue
            fn = fns[role]
            if fn is not None:
                cell.value = fn(cell.value)


def _anonymizeXlsx(data: bytes, state: _AnonState) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(data))
    for wsName in wb.sheetnames:
        _processSheet(
            wb[wsName],
            facilityFn=lambda v: state.mapFacility(str(v)),
            unitFn=lambda v: state.mapUnit(str(v)),
            latFn=lambda v: _offsetCoord(v, state.latOffset),
            lonFn=lambda v: _offsetCoord(v, state.lonOffset),
        )
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _deanonymizeXlsx(data: bytes, revFacility: dict, revUnit: dict,
                     latOffset: float, lonOffset: float) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(data))
    for wsName in wb.sheetnames:
        _processSheet(
            wb[wsName],
            facilityFn=lambda v: revFacility.get(str(v), v),
            unitFn=lambda v: revUnit.get(str(v), v),
            latFn=lambda v: _offsetCoord(v, -latOffset),
            lonFn=lambda v: _offsetCoord(v, -lonOffset),
        )
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _offsetCoord(val, offset: float):
    try:
        return round(float(val) + offset, 6)
    except (ValueError, TypeError):
        return val


def _studyZipName(stem: str) -> str:
    return f"{bf.STUDIES_DIR}/{stem}.xlsx"


def anonymizeBundle(zipPath: Path, keyPath: Path) -> None:
    """Anonymize a bundle zip in-place and write a key file to keyPath.

    Replaces Facility ID, Unit ID, Latitude, Longitude values in all study xlsx
    files with anonymized equivalents; renames study files; strips createdAt from
    metadata; anonymizes studyName in sim_config.
    """
    zipPath = Path(zipPath)
    keyPath = Path(keyPath)
    state   = _AnonState()

    studiesPrefix = bf.STUDIES_DIR + '/'

    with zipfile.ZipFile(zipPath, 'r') as zin:
        metadata  = json.loads(zin.read(bf.METADATA_FILE))
        simConfig = json.loads(zin.read(bf.SIM_CONFIG_FILE))

        # Pass 1: anonymize xlsx files, building mappings as we go
        anonStudyFiles: dict[str, bytes] = {}  # new zip path -> anonymized bytes
        for item in zin.infolist():
            name = item.filename
            if name.startswith(studiesPrefix) and name.endswith('.xlsx'):
                stem     = Path(name).stem
                anonStem = state.mapStudy(stem)
                anonData = _anonymizeXlsx(zin.read(name), state)
                anonStudyFiles[_studyZipName(anonStem)] = anonData

        # Anonymize sim_config studyName using the now-populated study map
        origStudyName = simConfig.get('studyName')
        if origStudyName:
            simConfig['studyName'] = state.studyMap.get(origStudyName,
                                                        state.mapStudy(origStudyName))

        # Strip creation timestamp from metadata
        metadata.pop('createdAt', None)

        # Pass 2: write anonymized zip to a temp file, then replace original
        tmpPath = zipPath.with_suffix('.anon_tmp')
        try:
            with zipfile.ZipFile(tmpPath, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                zout.writestr(bf.METADATA_FILE,   json.dumps(metadata,  indent=2))
                zout.writestr(bf.SIM_CONFIG_FILE, json.dumps(simConfig, indent=2))

                for anonName, anonData in anonStudyFiles.items():
                    zout.writestr(anonName, anonData)

                for item in zin.infolist():
                    name = item.filename
                    if (name in (bf.METADATA_FILE, bf.SIM_CONFIG_FILE) or
                            (name.startswith(studiesPrefix) and name.endswith('.xlsx'))):
                        continue
                    zout.writestr(item, zin.read(name))

            tmpPath.replace(zipPath)
        except Exception:
            tmpPath.unlink(missing_ok=True)
            raise

    keyPath.write_text(json.dumps(state.toKeyDict(), indent=2))
    logger.info(f"Anonymized bundle: {zipPath}")
    logger.info(f"Key file written:  {keyPath}")


def deanonymizeBundle(anonZipPath: Path, keyPath: Path, outputZipPath: Path) -> None:
    """Reverse an anonymized bundle using its key file.

    Restores original Facility IDs, Unit IDs, coordinates, study names, and
    xlsx filenames. Does not restore createdAt (not stored in key).
    """
    anonZipPath   = Path(anonZipPath)
    outputZipPath = Path(outputZipPath)
    key           = json.loads(Path(keyPath).read_text())

    latOffset  = key['coordinateOffset']['lat']
    lonOffset  = key['coordinateOffset']['lon']
    revStudy   = {v: k for k, v in key['studies'].items()}
    revFacility = {v: k for k, v in key['facilityIDs'].items()}
    revUnit    = {v: k for k, v in key['unitIDs'].items()}

    studiesPrefix = bf.STUDIES_DIR + '/'
    outputZipPath.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(anonZipPath, 'r') as zin, \
         zipfile.ZipFile(outputZipPath, 'w', compression=zipfile.ZIP_DEFLATED) as zout:

        simConfig = json.loads(zin.read(bf.SIM_CONFIG_FILE))
        anonStudyName = simConfig.get('studyName')
        if anonStudyName and anonStudyName in revStudy:
            simConfig['studyName'] = revStudy[anonStudyName]

        for item in zin.infolist():
            name = item.filename
            data = zin.read(name)

            if name.startswith(studiesPrefix) and name.endswith('.xlsx'):
                stem     = Path(name).stem
                origStem = revStudy.get(stem, stem)
                data     = _deanonymizeXlsx(data, revFacility, revUnit, latOffset, lonOffset)
                zout.writestr(_studyZipName(origStem), data)
            elif name == bf.SIM_CONFIG_FILE:
                zout.writestr(name, json.dumps(simConfig, indent=2))
            else:
                zout.writestr(item, data)

    logger.info(f"De-anonymized bundle written: {outputZipPath}")
